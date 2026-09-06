"""plan_7 step 5: bring the deposited files in, and decode what they actually are.

The pipeline route's acquisition is an nf-core/fetchngs run against SRA: 36.7 GB and ~5h15m on
study 22, 49.4 GB and ~17h on study 26. This one is an HTTP download and a decode. That difference
is the entire argument for the deposit route, so this module stays deliberately small: no pipeline
run, no Kubernetes, no notebook.

**Formats are sniffed from magic bytes, never from the extension.** This is not defensive
programming, it is what the real deposits require. Measured 2026-09-05 across the four studies this
project has run:

- ``GSE274331_TPMs_H2AS40-KD.xlsx`` is a TPM table in Excel. It is the only file in that deposit
  worth reproducing from, and study 22 spent five hours re-running the pipeline instead.
- ``GSE213770_DMR_DMB_TET2Neu.xls.gz`` gunzips to ``d0 cf 11 e0 a1 b1 1a e1``, the OLE2 compound
  document magic, so it is genuine legacy BIFF and not a mislabelled TSV.

Reading either as text produces mojibake that parses to zero rows, and a zero-row result is
indistinguishable from a paper that deposited nothing. Refusing by name is the honest outcome.
"""

from __future__ import annotations

import gzip
import io
import logging

logger = logging.getLogger("bioaf.deposit_acquisition")

# Default ceiling on one study's whole deposit download. A GEO supplementary directory can hold a
# `_RAW.tar` of every BAM in the study (835 MB on GSE274331, and that is a small one). `raw`-
# classified entries are never selectable, so this is the backstop rather than the first line.
DEFAULT_DOWNLOAD_CAP_BYTES = 2 * 1024**3

_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # legacy BIFF .xls (OLE2 compound document)
_ZIP_MAGIC = b"PK\x03\x04"  # .xlsx is a zip
_GZIP_MAGIC = b"\x1f\x8b"


class UnreadableDepositError(Exception):
    """A deposited file whose format we cannot turn into a table.

    Carries the FILENAME and the format, because this reaches a scientist at the C1 gate and
    "GSE213770_DMR.xls.gz is a legacy Excel file bioAF cannot read" is actionable where "could not
    parse" is not.
    """


class DepositTooLargeError(Exception):
    """The selected files exceed the download cap."""


def sniff_format(data: bytes) -> str:
    """What this payload actually is: ``gzip``, ``xlsx``, ``xls``, ``text`` or ``binary``.

    From the magic bytes. Depositors mislabel in both directions (a BIFF file named `.xls.gz`, a TSV
    named `.xls`), so the extension is a hint and the bytes are the evidence.
    """
    if not data:
        return "binary"
    if data.startswith(_GZIP_MAGIC):
        return "gzip"
    if data.startswith(_OLE2_MAGIC):
        return "xls"
    if data.startswith(_ZIP_MAGIC):
        return "xlsx"
    # A table is text. Decode strictly rather than with errors="replace": the whole point is to
    # notice binary rather than turn it into plausible-looking nonsense.
    try:
        data[:4096].decode("utf-8")
    except UnicodeDecodeError:
        return "binary"
    return "text"


def _xls_to_tsv(data: bytes, filename: str) -> str:
    """First worksheet of a legacy BIFF .xls as TSV.

    xlrd 2.x reads ONLY .xls: it deliberately dropped .xlsx support, which makes it the exact
    complement of openpyxl rather than an alternative to it. GSE213770's differential-methylation
    table is one of these, and until xlrd was added it was refused by name.
    """
    import xlrd

    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as exc:  # noqa: BLE001 - a corrupt workbook is a deposit problem, not a crash
        raise UnreadableDepositError(f"{filename} is a legacy Excel file that could not be opened: {exc}") from exc

    if not book.nsheets:
        raise UnreadableDepositError(f"{filename} is a legacy Excel file with no worksheets")
    sheet = book.sheet_by_index(0)

    lines: list[str] = []
    for r in range(sheet.nrows):
        cells = []
        for value in sheet.row_values(r):
            # xlrd returns every number as a float, so an integer count would render as "5.0" and
            # step 6 would measure the matrix as normalized rather than as counts. Integral floats
            # are written back as integers for exactly that reason.
            if isinstance(value, float) and value.is_integer():
                cells.append(str(int(value)))
            else:
                cells.append("" if value is None else str(value))
        if any(c.strip() for c in cells):
            lines.append("\t".join(cells))
    if not lines:
        raise UnreadableDepositError(f"{filename} is a legacy Excel file with no rows")
    return "\n".join(lines) + "\n"


def _xlsx_to_tsv(data: bytes, filename: str) -> str:
    """First worksheet of an .xlsx as TSV. openpyxl is already a dependency (requirements.txt)."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - a corrupt workbook is a deposit problem, not a crash
        raise UnreadableDepositError(f"{filename} is an Excel file that could not be opened: {exc}") from exc

    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        raise UnreadableDepositError(f"{filename} is an Excel file with no worksheets")

    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        cells = ["" if c is None else str(c) for c in row]
        if any(c.strip() for c in cells):
            lines.append("\t".join(cells))
    if not lines:
        raise UnreadableDepositError(f"{filename} is an Excel file with no rows")
    return "\n".join(lines) + "\n"


def decode_deposit(filename: str, raw: bytes) -> tuple[str, str]:
    """Turn a downloaded deposit file into (text, format), or refuse with a reason.

    Gunzips first, because a `.gz` wrapper hides the real format underneath and that is exactly how
    a BIFF spreadsheet came to be read as text.
    """
    data = raw
    fmt = sniff_format(data)
    if fmt == "gzip":
        try:
            data = gzip.decompress(data)
        except Exception as exc:  # noqa: BLE001
            raise UnreadableDepositError(f"{filename} could not be decompressed: {exc}") from exc
        fmt = sniff_format(data)

    if fmt == "text":
        return data.decode("utf-8", errors="replace"), "text"
    if fmt == "xlsx":
        return _xlsx_to_tsv(data, filename), "xlsx"
    if fmt == "xls":
        return _xls_to_tsv(data, filename), "xls"
    raise UnreadableDepositError(f"{filename} is not a readable table (format looks binary)")


def total_download_bytes(entries) -> int | None:
    """The summed size of a selection, or None when any size is unknown.

    None rather than a partial sum: a directory listing states no sizes, and treating an unknown as
    zero would let an unbounded file straight through the cap, which is the one thing the cap is
    for.
    """
    total = 0
    for e in entries or []:
        if e.size_bytes is None:
            return None
        total += int(e.size_bytes)
    return total


def check_download_cap(total: int | None, *, cap_bytes: int = DEFAULT_DOWNLOAD_CAP_BYTES) -> bool:
    """Whether the cap was checkable, raising when a known total exceeds it.

    Returns True when the total was known and fits, False when it could not be pre-checked. False is
    not a failure: most deposits list no sizes and are small. It is recorded so the gate can say the
    cap was not pre-checked rather than implying it passed.
    """
    if total is None:
        return False
    if total > cap_bytes:
        raise DepositTooLargeError(
            f"the selected deposit files total {total / 1024**3:.1f} GB, over the "
            f"{cap_bytes / 1024**3:.1f} GB limit for a deposit download"
        )
    return True
