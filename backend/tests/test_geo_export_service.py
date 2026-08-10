"""Tests for GeoExportService orchestrator."""

import zipfile
from io import BytesIO

import pytest
import pytest_asyncio
from app.services.auth_service import AuthService


@pytest_asyncio.fixture
async def comp_bio_user(session, admin_user):
    from app.models.user import User

    password_hash = AuthService.hash_password("compbiopass123")
    user = User(
        email="compbio@test.com",
        password_hash=password_hash,
        role_id=admin_user._test_role_map["comp_bio"],
        organization_id=admin_user.organization_id,
        status="active",
    )
    session.add(user)
    await session.flush()
    await session.commit()
    return user


@pytest_asyncio.fixture
async def export_experiment(client, admin_token, session, admin_user):
    """Create experiment with samples, batch, pipeline run for GEO export."""
    from app.models.sample_batch import SampleBatch
    from app.models.pipeline_run import PipelineRun
    from app.models.sample import Sample

    org_id = admin_user.organization_id

    # Create experiment
    resp = await client.post(
        "/api/experiments",
        json={"name": "GEO Export Test", "description": "Test experiment for GEO export."},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    exp_id = resp.json()["id"]

    # Create batch
    batch = SampleBatch(
        experiment_id=exp_id,
        name="Batch1",
    )
    session.add(batch)
    await session.flush()

    # Create samples
    s1 = Sample(
        experiment_id=exp_id,
        sample_batch_id=batch.id,
        external_id="SAMPLE_001",
        organism="Homo sapiens",
        tissue_type="blood",
        molecule_type="total RNA",
        library_prep_method="10x Chromium 3' v3.1",
        library_layout="paired",
        treatment_condition="Vehicle",
        qc_status="pass",
    )
    s2 = Sample(
        experiment_id=exp_id,
        sample_batch_id=batch.id,
        external_id="SAMPLE_002",
        organism="Homo sapiens",
        tissue_type="liver",
        molecule_type="total RNA",
        library_prep_method="10x Chromium 3' v3.1",
        library_layout="paired",
        treatment_condition="Drug X",
        qc_status="fail",
    )
    session.add_all([s1, s2])
    await session.flush()

    # Create pipeline run
    run = PipelineRun(
        organization_id=org_id,
        experiment_id=exp_id,
        pipeline_name="nf-core/scrnaseq",
        pipeline_version="2.7.0",
        status="completed",
        reference_genome="GRCh38",
        alignment_algorithm="STARsolo",
    )
    session.add(run)
    await session.flush()
    await session.commit()

    return {
        "experiment_id": exp_id,
        "org_id": org_id,
        "run_id": run.id,
        "sample_ids": [s1.id, s2.id],
    }


@pytest.mark.asyncio
async def test_validate_returns_report(session, export_experiment):
    from app.services.geo.geo_export_service import GeoExportService

    report = await GeoExportService.validate(
        session,
        export_experiment["experiment_id"],
        export_experiment["org_id"],
        pipeline_run_id=export_experiment["run_id"],
    )

    assert report.experiment_id == export_experiment["experiment_id"]
    assert report.pipeline_run_id == export_experiment["run_id"]
    assert report.summary.total_fields > 0
    # With exclude_failed, only 1 sample should be validated
    assert len(report.sample_validations) == 1


@pytest.mark.asyncio
async def test_export_returns_zip(session, export_experiment):
    from app.services.geo.geo_export_service import GeoExportService

    zip_bytes, filename = await GeoExportService.export(
        session,
        export_experiment["experiment_id"],
        export_experiment["org_id"],
        pipeline_run_id=export_experiment["run_id"],
    )

    assert filename.startswith("geo_export_")
    assert filename.endswith(".zip")

    # Verify ZIP contents
    zf = zipfile.ZipFile(BytesIO(zip_bytes))
    names = zf.namelist()
    assert any(n.endswith(".xlsx") for n in names)
    assert "md5_checksums.txt" in names
    assert "validation_report.json" in names
    assert "validation_report.txt" in names
    assert "README.txt" in names


@pytest.mark.asyncio
async def test_qc_filter_excludes_failed(session, export_experiment):
    from app.services.geo.geo_export_service import GeoExportService

    # With filter (default) - should exclude the "fail" sample
    report = await GeoExportService.validate(
        session,
        export_experiment["experiment_id"],
        export_experiment["org_id"],
        qc_status_filter="exclude_failed",
    )
    assert len(report.sample_validations) == 1
    assert report.sample_validations[0].sample_name == "SAMPLE_001"

    # Without filter - should include both
    report_all = await GeoExportService.validate(
        session,
        export_experiment["experiment_id"],
        export_experiment["org_id"],
        qc_status_filter="include_all",
    )
    assert len(report_all.sample_validations) == 2


@pytest.mark.asyncio
async def test_auto_selects_pipeline_run(session, export_experiment):
    from app.services.geo.geo_export_service import GeoExportService

    # Don't specify pipeline_run_id: should auto-select the completed run
    report = await GeoExportService.validate(
        session,
        export_experiment["experiment_id"],
        export_experiment["org_id"],
    )
    assert report.pipeline_run_id == export_experiment["run_id"]


@pytest.mark.asyncio
async def test_nonexistent_experiment_raises(session):
    from app.exceptions import NotFoundError
    from app.services.geo.geo_export_service import GeoExportService

    with pytest.raises(NotFoundError, match="Experiment not found"):
        await GeoExportService.validate(session, 99999, 99999)


@pytest.mark.asyncio
async def test_export_excel_readable(session, export_experiment):
    """Verify the Excel file inside the ZIP is actually openable."""
    from openpyxl import load_workbook

    from app.services.geo.geo_export_service import GeoExportService

    zip_bytes, _ = await GeoExportService.export(
        session,
        export_experiment["experiment_id"],
        export_experiment["org_id"],
    )

    zf = zipfile.ZipFile(BytesIO(zip_bytes))
    xlsx_name = [n for n in zf.namelist() if n.endswith(".xlsx")][0]
    xlsx_data = zf.read(xlsx_name)

    wb = load_workbook(BytesIO(xlsx_data))
    assert wb.sheetnames == ["SERIES", "SAMPLES", "PROTOCOLS"]
    # Data should be present
    ws = wb["SAMPLES"]
    assert ws.max_row >= 2  # Header + at least 1 sample


def test_validation_report_uses_geo_tag_value_convention():
    """The unvalidated-values section must follow GEO's `tag: value` idiom.

    GEO's SOFT format and the sample characteristics field express metadata as
    colon-separated `tag: value` pairs (e.g. `tissue: liver`), and the sibling
    MISSING REQUIRED lines in this same report already use `{geo_column}: ...`.

    A comma separator is actively wrong here: GEO values routinely contain commas
    (`tissue: liver, frozen`), so `column: value, message` cannot be read back
    unambiguously. The advisory message therefore goes in parentheses.
    """
    from app.services.geo.geo_export_service import _format_validation_report
    from app.services.geo.validation import (
        FieldValidation,
        FileManifestStatus,
        SampleValidation,
        ValidationReport,
        ValidationSummary,
    )

    field = FieldValidation(
        geo_column="tissue",
        value="liver, frozen",
        status="populated_unvalidated",
        message="not in controlled vocabulary",
    )
    report = ValidationReport(
        experiment_id=1,
        pipeline_run_id=None,
        summary=ValidationSummary(
            total_fields=1,
            complete=0,
            populated_unvalidated=1,
            missing_required=0,
            missing_recommended=0,
        ),
        series_fields=[field],
        protocol_fields=[],
        sample_validations=[SampleValidation(sample_id=1, sample_name="S1", fields=[field])],
        file_manifest=FileManifestStatus(total_files=0, files_with_checksums=0, files_missing_checksums=0, files=[]),
    )

    text = _format_validation_report(report)

    assert "[SERIES/PROTOCOL] tissue: liver, frozen (not in controlled vocabulary)" in text
    assert "[SAMPLE S1] tissue: liver, frozen (not in controlled vocabulary)" in text
    # No em-dash anywhere in generated output.
    assert "—" not in text
